# -*- coding: utf-8 -*-
# @time    : 2023/9/21 10:03
# @author  : w-xin
# @file    : xlsx_transform_demo_v2.py
# @software: PyCharm

"""
"""
import os
import warnings

import numpy as np
import openpyxl
import pandas as pd
from tqdm import tqdm


def merge_xlsx_df(path):

    all_dfs = []
    file_list_1th = os.listdir(path)
    for file1 in file_list_1th:
        if '.' not in file1:
            print(f'-- fonder {file1} start --')
            path_2th = os.path.join(path, file1)
            file_list_2th = os.listdir(path_2th)
            for file2 in tqdm(file_list_2th):
                if '.xlsx' in file2:
                    # 打开一个XLSX文件
                    workbook = openpyxl.load_workbook(f'./{path_2th}/{file2}')

                    # 选择一个工作表
                    sheet = workbook.active

                    data = []
                    # 遍历行和列
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                        row_data = [cell.value for cell in row]
                        data.append(row_data)
                    # 将数据列表转换为DataFrame
                    detail = pd.DataFrame(data)

                    for col in detail.columns:
                        detail[col] = detail[col].map(lambda x: str(x).replace('None', 'nan'))

                    for idx in detail.index:
                        if '户名' in detail.iloc[idx, 0]:
                            username = detail.iloc[idx, 0].replace('户名：', '')
                        elif '账号' in detail.iloc[idx, 0] or '卡号' in detail.iloc[idx, 0]:
                            account = detail.iloc[idx, 0].replace('账号/卡号：', '')
                        elif '序号' in detail.iloc[idx, 0]:
                            # 去除用户信息 提取流水表格
                            stream = detail.iloc[idx:, :].reset_index(drop=True)
                            stream.columns = stream.iloc[0, :]
                            stream.drop(labels=[0], axis=0, inplace=True)
                            stream.reset_index(drop=True, inplace=True)
                    stream = stream[~stream['序号'].str.contains('\n')]

                    # 标记跨行数
                    row_take_num = np.array([])
                    temp = 1
                    for r in range(len(stream.iloc[:, 0].values)):
                        try:
                            if str(stream.iloc[r + 1, 0]) != 'nan':
                                row_take_num = np.append(row_take_num, np.array(temp))
                                temp = 1
                            else:
                                temp += 1
                        except:
                            row_take_num = np.append(row_take_num, np.array(temp))

                    # stream_copy = stream.copy()

                    # 合并行
                    stream_split_list = []
                    for idx, row_num in enumerate(row_take_num):
                        row_num = int(row_num)
                        temt_df = stream.iloc[: row_num, :]
                        if len(temt_df.index) == 1:
                            temt_df = pd.DataFrame(temt_df.iloc[0, :]).T
                            temt_df.insert(loc=1, column='户名', value=username)
                            temt_df.insert(loc=2, column='交易卡号', value=account)
                            temt_df.index = [idx]
                            stream_split_list += [temt_df]
                        else:
                            for col_num in range(1, len(stream.columns)):
                                if str(temt_df.iloc[1, col_num]) == 'nan':
                                    continue
                                elif '金额' in temt_df.columns[col_num] or '余额' in temt_df.columns[col_num]:
                                    nan_count = list(temt_df.iloc[:, col_num].isnull()).count(True)
                                    valid_row = row_num - nan_count
                                    for row in range(1, valid_row):
                                        if str(temt_df.iloc[row, col_num]) == 'nan':
                                            break
                                        else:
                                            if '.' in str(temt_df.iloc[0, col_num]):
                                                temt_df.iloc[0, col_num] = str(temt_df.iloc[0, col_num]) + str(temt_df.iloc[row, col_num])
                                            else:
                                                temt_df.iloc[0, col_num] = str(temt_df.iloc[0, col_num]) + '.' + str(temt_df.iloc[row, col_num])
                                else:
                                    nan_count = list(temt_df.iloc[:, col_num].isnull()).count(True)
                                    valid_row = row_num - nan_count
                                    for row in range(1, valid_row):
                                        if str(temt_df.iloc[row, col_num]) == 'nan':
                                            break
                                        else:
                                            temt_df.iloc[0, col_num] = str(temt_df.iloc[0, col_num]) + str(temt_df.iloc[row, col_num])
                            temt_df = pd.DataFrame(temt_df.iloc[0, :]).T
                            temt_df.insert(loc=1, column='户名', value=username)
                            temt_df.insert(loc=2, column='交易卡号', value=account)
                            temt_df.index = [idx]
                            stream_split_list += [temt_df]
                        stream = stream.iloc[row_num:, :].reset_index(drop=True)
                    all_dfs += [pd.concat(stream_split_list)]

    df = pd.concat(all_dfs)

    return df


if __name__ == '__main__':

    warnings.filterwarnings('ignore')
    pd.set_option('display.max_columns', 1000)
    pd.set_option('display.width', 1000)

    # detail2 = pd.read_excel('./2明细.xlsx')
    # detail10 = pd.read_excel('./10明细列表【2023-08-14】.xlsx')

    #
    # # 打开一个XLSX文件
    # workbook = openpyxl.load_workbook('./data/11明细列表【2023-08-14】.xlsx')
    #
    # # 选择一个工作表
    # sheet = workbook.active
    #
    # data = []
    # # 遍历行和列
    # for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    #     row_data = [cell.value for cell in row]
    #     data.append(row_data)
    #
    # # 将数据列表转换为DataFrame
    # detail = pd.DataFrame(data)
    # for col in detail.columns:
    #     detail[col] = detail[col].map(lambda x: str(x).replace('None', 'nan'))
    # detail = detail.replace('None', 'nan')
    #
    # for idx in detail.index:
    #     if '户名' in detail.iloc[idx, 0]:
    #         username = detail.iloc[idx, 0].replace('户名：', '')
    #     elif '账号' in detail.iloc[idx, 0] or '卡号' in detail.iloc[idx, 0]:
    #         account = detail.iloc[idx, 0].replace('账号/卡号：', '')
    #     elif '序号' in detail.iloc[idx, 0]:
    #         # 去除用户信息 提取流水表格
    #         stream = detail.iloc[idx:, :].reset_index(drop=True)
    #         stream.columns = stream.iloc[0, :]
    #         stream.drop(labels=[0], axis=0, inplace=True)
    #         stream.reset_index(drop=True, inplace=True)
    #
    #
    # stream = stream[~stream['序号'].str.contains('\n')]
    # # 标记跨行数
    # row_take_num = np.array([])
    # temp = 1
    # all_dfs = []
    # for r in range(len(stream.iloc[:, 0].values)):
    #     try:
    #         if str(stream.iloc[r + 1, 0]) != 'nan':
    #             row_take_num = np.append(row_take_num, np.array(temp))
    #             temp = 1
    #         else:
    #             temp += 1
    #     except:
    #         row_take_num = np.append(row_take_num, np.array(temp))
    #
    # stream_copy = stream.copy()
    # # 合并行
    # stream_split_list = []
    # for idx, row_num in enumerate(row_take_num):
    #     row_num = int(row_num)
    #     temt_df = stream.iloc[: row_num, :]
    #     if len(temt_df.index) == 1:
    #         temt_df = pd.DataFrame(temt_df.iloc[0, :]).T
    #         temt_df.insert(loc=1, column='户名', value=username)
    #         temt_df.insert(loc=2, column='交易卡号', value=account)
    #         temt_df.index = [idx]
    #         stream_split_list += [temt_df]
    #     else:
    #         for col_num in range(1, len(stream.columns)):
    #             if str(temt_df.iloc[1, col_num]) == 'nan':
    #                 continue
    #             elif '金额' in temt_df.columns[col_num] or '余额' in temt_df.columns[col_num]:
    #                 nan_count = list(temt_df.iloc[:, col_num].isnull()).count(True)
    #                 valid_row = row_num - nan_count
    #                 for row in range(1, valid_row):
    #                     if str(temt_df.iloc[row, col_num]) == 'nan':
    #                         break
    #                     else:
    #                         temt_df.iloc[0, col_num] = str(temt_df.iloc[0, col_num]) + '.' + str(
    #                             temt_df.iloc[row, col_num])
    #             else:
    #                 nan_count = list(temt_df.iloc[:, col_num].isnull()).count(True)
    #                 valid_row = row_num - nan_count
    #                 for row in range(1, valid_row):
    #                     if str(temt_df.iloc[row, col_num]) == 'nan':
    #                         break
    #                     else:
    #                         temt_df.iloc[0, col_num] = str(temt_df.iloc[0, col_num]) + str(temt_df.iloc[row, col_num])
    #         temt_df = pd.DataFrame(temt_df.iloc[0, :]).T
    #         temt_df.insert(loc=1, column='户名', value=username)
    #         temt_df.insert(loc=2, column='交易卡号', value=account)
    #         temt_df.index = [idx]
    #         stream_split_list += [temt_df]
    #     stream = stream.iloc[row_num:, :].reset_index(drop=True)
    # all_dfs += [pd.concat(stream_split_list)]
    #     # stream[stream['序号'] == '1115\n1116']

    path = './'

    all_data = merge_xlsx_df(path)
    # '14534.0.64'
    # all_data.iloc[909, 8].replace('.0', '0')